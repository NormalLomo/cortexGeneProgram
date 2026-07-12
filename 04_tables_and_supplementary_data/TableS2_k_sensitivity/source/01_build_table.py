#!/usr/bin/env python3
"""
build_tableS4.py
Build TableS4_k_sensitivity.xlsx from TableS4_k_sensitivity.tsv
Input:  supplementary/TableS4_k_sensitivity.tsv
Output: supplementary/TableS4_k_sensitivity.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPP_DIR = os.path.join(os.path.dirname(__file__), "")
TSV_PATH = os.path.join(SUPP_DIR, "TableS4_k_sensitivity.tsv")
XLSX_PATH = os.path.join(SUPP_DIR, "TableS4_k_sensitivity.xlsx")

# ── read TSV ──────────────────────────────────────────────────────────────────
rows = []
with open(TSV_PATH, encoding="utf-8") as f:
    for line in f:
        rows.append(line.rstrip("\n").split("\t"))

header = rows[0]
data   = rows[1:]

# ── workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Table S4 K-sensitivity"

# ── legend block (rows 1-5) ────────────────────────────────────────────────────
legend_lines = [
    "Table S4. K-sensitivity of the program backbone (K = 40/50/65/70/80 vs K = 60).",
    "",
    "Consensus cNMF was re-run at K = 40, 50, 65, 70 and 80 and the resulting programs",
    "were matched back to the K = 60 backbone by cosine similarity. Spearman rho and",
    "Pearson r are the cross-area region-variability rank correlations between the",
    "alternative-K matched programs and the K = 60 reference programs (Fig. S1).",
    "Top driver subclass: the #1 subclass ranked by cross-area region effect size (eta2)",
    "across all matched programs at that K. Top region-variable program subclass: the",
    "dominant cell subclass of the single most region-variable matched program at that K.",
    "K = 60 (reference) values are shown for comparison; 8 of 54 programs are",
    "region-variable and 39 of 54 programs have spatial top cell-type matching their",
    "single-nucleus dominant subclass at K = 60 (Fig. S1, Fig. S7; Methods).",
    "Sources: Fig. S1 legend and Methods (Choice of K) in the main figure documentation.",
]
for i, line in enumerate(legend_lines, start=1):
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = Font(italic=(i == 1), bold=(i == 1), size=10)
    cell.alignment = Alignment(wrap_text=True)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
for i in range(2, len(legend_lines) + 1):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(header))

HEADER_ROW = len(legend_lines) + 2  # blank row then header

# ── styles ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
REF_FILL    = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL    = PatternFill("solid", fgColor="EBF5FB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(size=10)
REF_FONT    = Font(size=10, italic=True)
BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE,  bottom=BORDER_SIDE)

# ── human-readable column headers ────────────────────────────────────────────
DISPLAY_HEADERS = [
    "K (factorization rank)",
    "Spearman ρ vs K = 60",
    "Pearson r vs K = 60",
    "Significance",
    "#1 driver subclass",
    "Top region-variable program subclass",
    "Conclusion stable?",
]

# ── header row ────────────────────────────────────────────────────────────────
for col_idx, col_name in enumerate(DISPLAY_HEADERS, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
    cell.font    = HEADER_FONT
    cell.fill    = HEADER_FILL
    cell.border  = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── data rows ─────────────────────────────────────────────────────────────────
for row_idx, row in enumerate(data, start=HEADER_ROW + 1):
    is_ref = "reference" in row[0].lower()
    fill   = REF_FILL if is_ref else (ALT_FILL if row_idx % 2 == 0 else PatternFill())
    font   = REF_FONT if is_ref else BODY_FONT
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = font
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_idx in (5, 6):  # subclass columns → left-align
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [20, 22, 20, 22, 28, 52, 20]
for col_idx, width in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── row heights ───────────────────────────────────────────────────────────────
ws.row_dimensions[HEADER_ROW].height = 32
for row_idx in range(HEADER_ROW + 1, HEADER_ROW + 1 + len(data)):
    ws.row_dimensions[row_idx].height = 28

# ── freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

wb.save(XLSX_PATH)
print(f"Saved: {XLSX_PATH}")
